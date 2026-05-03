from pathlib import Path

import pandas as pd
import xlrd
from openpyxl import load_workbook


SUPPORTED_SUFFIXES = {".xlsx", ".xls"}


def load_first_sheet(path: Path) -> tuple[pd.DataFrame, str]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("仅支持导入 .xlsx 或 .xls 格式的 Excel 文件")
    try:
        if suffix == ".xls":
            workbook = pd.read_excel(path, sheet_name=None, engine="xlrd")
            sheet_name = next(iter(workbook))
            return workbook[sheet_name], sheet_name
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = workbook.sheetnames[0]
        workbook.close()
        frame = pd.read_excel(path, sheet_name=sheet_name)
        return frame, sheet_name
    except Exception as exc:
        raise ValueError(f"无法读取该 Excel 文件：{exc}") from exc


def export_dataframe(frame: pd.DataFrame, target: Path) -> None:
    frame.to_excel(target, index=False)
